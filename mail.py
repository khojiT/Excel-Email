#It's khoji <a.khoji2001@gmail.com>
# * in this code means that you should put your individual information

#pip install openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import openpyxl
import os
#readxlsx function is used to read excels
def readxlsx(name,path,sheet,column,input_list):
    os.chdir(path)
    ex = openpyxl.load_workbook(name)
    sheet = ex[sheet]
    for i in range(1, sheet.max_row + 1): input_list += [sheet.cell(row=i, column=column).value]
    return input_list
l2 =[]
l1 =[]
#There is no limit to the number of columns you can read
readxlsx('*name.xlsx','*C:\\Users\\persian\\Desktop','*Sheet1',*2,l2)
#open_file function is used to read files that you want to send
def open_file(path):
    with open(path,'rb') as f:
        ext = os.path.splitext(f.name)
        name_of_file = (ext[0].split('\\')[-1])
        app = MIMEApplication(f.read())
    return app,name_of_file,ext[1]
#Select the files which you want to send(no limit)
res1 = open_file(r'*C:\\Users\\persian\\Desktop\\name')

html = """*\
<html>
  <head></head>
  <body>
    <h1>Hi!<br>
       How are you bro?<br>
       Here is the <a href="http://www.python.org">link</a> you wanted.
    </h1>
  </body>
</html>
"""
#your gmail account as sender
#with that account got to https://myaccount.google.com/lesssecureapps?pli=1&rapt=AEjHL4OORDVc9fBKaPH0n8_4UEBlE5NB8K_16ZtEkdsPpeDaiudKCcpxtwtkFbwBJiGexLcjr04-CtSMS5zq9me7xEf1EYeGVQ and turn it on
sender ="*email.email@gmail.com"
password="*password"
#header and attachment
def add_head_att(res):
    res[0].add_header('Content-Disposition', 'attachment', filename=res[1]+res[2])
    message.attach(res[0])
for receiver,name in zip(l2,l1):
    name = ''.join(name)
    print(name)
    message = MIMEMultipart()
    message['From'] = sender
    message['to'] = receiver
    message['subject'] = "*It's the subject of my email "
    body = f'*Dear {name} \nwhat\'s ubody?'
    message.attach(MIMEText(body, 'plain'))
    message.attach(MIMEText(html,'html'))
    add_head_att(res1)
    mail = smtplib.SMTP('smtp.gmail.com', 587)
    mail.ehlo()
    mail.starttls()
    mail.login(sender, password)
    text = message.as_string()
    mail.sendmail(sender,receiver,text)
